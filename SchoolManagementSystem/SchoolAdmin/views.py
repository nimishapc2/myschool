from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import TeacherCreationForm, TeacherProfileForm, StudentCreationForm,StudentForm
from django.shortcuts import get_object_or_404
from SchoolAdmin.models import Teacher, Subject, Division, TeacherAssignment, SchoolClass, Student, Attendance, Announcement
from django.contrib import messages
from django.db.models import Q
from datetime import date

# Create your views here.

def dashboard(request):
    print("Is Authenticated:", request.user.is_authenticated)
    return render(request, "admin_dashboard.html")


def is_admin(user):
    return user.groups.filter(name='Admin').exists()

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    return render(request, 'SchoolAdmin/dashboard.html')


# Add Teacher
@login_required
def add_teacher(request):
    if request.user.role != "ADMIN":
        return redirect("login")

    if request.method == "POST":
        user_form = TeacherCreationForm(request.POST)
        profile_form = TeacherProfileForm(request.POST)
        print(user_form.errors)
        print(profile_form.errors)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save(commit=False)
            user.role = "TEACHER"
            user.save()

            teacher = profile_form.save(commit=False)
            teacher.user = user
            teacher.save()

            return redirect("admin_dashboard")
    else:
        user_form = TeacherCreationForm()
        profile_form = TeacherProfileForm()

    return render(request, "add_teacher.html", {
        "user_form": user_form,
        "profile_form": profile_form
    })


# View All teachers

@login_required
def view_teachers(request):
    if request.user.role != "ADMIN":
        return redirect("login")

    # DELETE
    if request.method == "POST" and "delete_id" in request.POST:
        teacher = get_object_or_404(Teacher, id=request.POST.get("delete_id"))
        teacher.user.delete()
        return redirect("SchoolAdmin:view_teachers")

    # UPDATE
    if request.method == "POST" and "edit_id" in request.POST:
        teacher = get_object_or_404(Teacher, id=request.POST.get("edit_id"))

        teacher.name = request.POST.get("name")
        teacher.subject = request.POST.get("subject")
        teacher.qualification = request.POST.get("qualification")
        teacher.contact_number = request.POST.get("contact_number")
        teacher.address = request.POST.get("address")
        teacher.save()

        return redirect("SchoolAdmin:view_teachers")

    teachers = Teacher.objects.select_related("user").all()

    return render(request, "view_teachers.html", {"teachers": teachers})


# Subject Assignment

def subject_assignments(request):

    teachers = Teacher.objects.all()
    subjects = Subject.objects.select_related("school_class").all()
    divisions = Division.objects.select_related("school_class").all()
    assignments = TeacherAssignment.objects.select_related(
        "teacher", "subject", "division", "division__school_class"
    ).all()

    # ADD Assignment
    if request.method == "POST":
        teacher_id = request.POST.get("teacher")
        subject_id = request.POST.get("subject")
        division_id = request.POST.get("division")

        try:
            TeacherAssignment.objects.create(
                teacher_id=teacher_id,
                subject_id=subject_id,
                division_id=division_id
            )
            messages.success(request, "Assignment added successfully.")
        except:
            messages.error(request, "Assignment already exists.")

        return redirect("SchoolAdmin:subject_assignments")

    context = {
        "teachers": teachers,
        "subjects": subjects,
        "divisions": divisions,
        "assignments": assignments,
    }

    return render(request, "subject_assignments.html", context)

# teacher assignment view

from django.db.models import Q

def teacher_assignment_view(request):

    teachers = Teacher.objects.all()
    subjects = Subject.objects.all()
    divisions = Division.objects.select_related("school_class").all()

    selected_teacher = None
    assignments = None

    # Search teacher
    if request.GET.get("teacher"):
        teacher_id = request.GET.get("teacher")
        selected_teacher = Teacher.objects.get(id=teacher_id)
        assignments = TeacherAssignment.objects.filter(
            teacher=selected_teacher
        ).select_related("subject", "division", "division__school_class")

    # Inline Update
    if request.method == "POST":
        assignment_id = request.POST.get("assignment_id")
        assignment = TeacherAssignment.objects.get(id=assignment_id)

        if "delete" in request.POST:
            assignment.delete()
        else:
            assignment.subject_id = request.POST.get("subject")
            assignment.division_id = request.POST.get("division")
            assignment.save()

        return redirect(
            f"{request.path}?teacher={assignment.teacher.id}"
        )

    context = {
        "teachers": teachers,
        "subjects": subjects,
        "divisions": divisions,
        "assignments": assignments,
        "selected_teacher": selected_teacher,
    }

    return render(request, "teacher_assignment_view.html", context)

# Add student

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction

@login_required
def add_student(request):

    if request.user.role != "ADMIN":
        messages.error(request, "You are not authorized to access this page.")
        return redirect("login")

    if request.method == "POST":
        user_form = StudentCreationForm(request.POST)
        student_form = StudentForm(request.POST)
        print("USER VALID:", user_form.is_valid())
        print("STUDENT VALID:", student_form.is_valid())
        print("USER ERRORS:", user_form.errors)
        print("STUDENT ERRORS:", student_form.errors)
        if user_form.is_valid() and student_form.is_valid():
            try:
                with transaction.atomic():

                    # Create User
                    user = user_form.save(commit=False)
                    user.role = "STUDENT"
                    user.save()

                    # Create Student Profile
                    student = student_form.save(commit=False)
                    student.user = user
                    student.save()

                    messages.success(request, "Student added successfully!")
                    return redirect("admin_dashboard")

            except Exception as e:
                messages.error(request, f"Error occurred: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")

    else:
        user_form = StudentCreationForm()
        student_form = StudentForm()

    return render(request, "add_student.html", {
        "user_form": user_form,
        "student_form": student_form
    })


# View Students

def view_students(request):

    students = None  # Do not load initially

    classes = SchoolClass.objects.all()
    divisions = Division.objects.select_related("school_class").all()

    name_query = request.GET.get("name")
    class_query = request.GET.get("class")
    division_query = request.GET.get("division")

    # 🔍 SEARCH
    if name_query or class_query or division_query:

        students = Student.objects.select_related(
            "division", "division__school_class"
        ).all()

        if name_query:
            students = students.filter(name__icontains=name_query)

        if class_query:
            students = students.filter(
                division__school_class_id=class_query
            )

        if division_query:
            students = students.filter(
                division_id=division_query
            )

    # ✏ UPDATE / DELETE
    if request.method == "POST":

        student_id = request.POST.get("student_id")
        student = get_object_or_404(Student, id=student_id)

        # DELETE
        if "delete" in request.POST:
            student.delete()
            messages.success(request, "Student deleted successfully.")
            return redirect(request.get_full_path())

        # UPDATE
        if "update" in request.POST:
            student.name = request.POST.get("name")
            student.guardian_name = request.POST.get("guardian_name")
            student.contact_number = request.POST.get("contact_number")
            student.address = request.POST.get("address")
            student.division_id = request.POST.get("division")

            student.save()
            messages.success(request, "Student updated successfully.")
            return redirect(request.get_full_path())

    context = {
        "students": students,
        "classes": classes,
        "divisions": divisions,
    }

    return render(request, "view_students.html", context)

# Attendance Posting



def mark_attendance(request):

    classes = SchoolClass.objects.all()
    divisions = Division.objects.select_related("school_class").all()
    students = None

    class_id = request.GET.get("class")
    division_id = request.GET.get("division")

    # 🔍 Filter students
    if class_id and division_id:
        students = Student.objects.filter(
            division_id=division_id,
            division__school_class_id=class_id
        )

    # 💾 Save attendance
    if request.method == "POST":
        division_id = request.POST.get("division")
        class_id = request.POST.get("class")
        today = date.today()

        students = Student.objects.filter(
            division_id=division_id,
            division__school_class_id=class_id
        )

        for student in students:
            status = request.POST.get(f"student_{student.id}") == "on"

            Attendance.objects.update_or_create(
                student=student,
                date=today,
                defaults={"status": status}
            )

        messages.success(request, "Attendance saved successfully.")
        return redirect(request.get_full_path())

    context = {
        "classes": classes,
        "divisions": divisions,
        "students": students
    }

    return render(request, "mark_attendance.html", context)

# Attendance Report

from django.db.models import Count, Q
from datetime import datetime



def attendance_report(request):

    classes = SchoolClass.objects.all()
    divisions = Division.objects.select_related("school_class").all()

    records = None
    total_students = 0
    present_count = 0
    absent_count = 0

    class_id = request.GET.get("class")
    division_id = request.GET.get("division")
    date_str = request.GET.get("date")

    if class_id and division_id and date_str:

        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()

        records = Attendance.objects.filter(
            student__division_id=division_id,
            student__division__school_class_id=class_id,
            date=selected_date
        ).select_related("student")

        total_students = Student.objects.filter(
            division_id=division_id,
            division__school_class_id=class_id
        ).count()

        present_count = records.filter(status=True).count()
        absent_count = total_students - present_count

    context = {
        "classes": classes,
        "divisions": divisions,
        "records": records,
        "total_students": total_students,
        "present_count": present_count,
        "absent_count": absent_count,
    }

    return render(request, "attendance_report.html", context)

# Monthly Attendance

from django.db.models import Count, Q

import calendar



import calendar


def monthly_attendance(request):
    classes = SchoolClass.objects.all()
    divisions = Division.objects.all()

    report = []
    total_working_days = 0

    class_id = request.GET.get('class_id')
    division_id = request.GET.get('division_id')
    month = request.GET.get('month')

    if class_id and division_id and month:

        year, month_number = map(int, month.split('-'))

        # Filter students properly
        students = Student.objects.filter(
            division_id=division_id,
            division__school_class_id=class_id
        )

        # Total days in month
        total_working_days = calendar.monthrange(year, month_number)[1]

        for student in students:
            present_count = Attendance.objects.filter(
                student=student,
                date__year=year,
                date__month=month_number,
                status=True
            ).count()

            absent_count = total_working_days - present_count

            percentage = 0
            if total_working_days > 0:
                percentage = round((present_count / total_working_days) * 100, 2)

            report.append({
                'student': student,
                'present': present_count,
                'absent': absent_count,
                'percentage': percentage
            })

    context = {
        'classes': classes,
        'divisions': divisions,
        'report': report,
        'total_working_days': total_working_days
    }

    return render(request, 'monthly_attendance.html', context)


# Monthly Attendance Report Excel
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar

from .models import Student, Attendance, Division


def monthly_attendance_excel(request):
    class_id = request.GET.get("class_id")
    division_id = request.GET.get("division_id")
    month = request.GET.get("month")  # format: 2026-02

    if not (class_id and division_id and month):
        return HttpResponse("Missing required parameters")

    year, month_number = map(int, month.split("-"))

    # Get students of selected class + division
    students = Student.objects.filter(
        division_id=division_id,
        division__school_class_id=class_id
    )

    total_days = calendar.monthrange(year, month_number)[1]

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Monthly Attendance Report - {month}"
    ws["A1"].font = Font(size=14, bold=True)

    # Headers
    headers = ["Sl No", "Student Name", "Present Days", "Absent Days", "Attendance %"]
    ws.append(headers)

    for cell in ws[2]:
        cell.font = Font(bold=True)

    # Add student data
    for index, student in enumerate(students, start=1):

        present_days = Attendance.objects.filter(
            student=student,
            date__year=year,
            date__month=month_number,
            status=True
        ).count()

        absent_days = total_days - present_days

        percentage = round((present_days / total_days) * 100, 2) if total_days > 0 else 0

        ws.append([
            index,
            student.name,
            present_days,
            absent_days,
            f"{percentage}%"
        ])

    # Auto column width
    for col in range(1, 6):
        max_length = 0
        column_letter = get_column_letter(col)

        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="monthly_attendance_{month}.xlsx"'

    wb.save(response)
    return response

# --------------------
# POST ANNOUNCEMENT
# --------------------
# @login_required
def add_announcement(request):
    if request.method == 'POST':
        title = request.POST['title']
        message = request.POST['message']

        Announcement.objects.create(title=title, message=message)
        return redirect('admin_dashboard')

    return render(request, 'add_announcement.html')

# --------------------
# ANNOUNCEMENT LIST
# --------------------


def announcement_list(request):
    announcements = Announcement.objects.all().order_by('-created_at')

    # 🔍 SEARCH (title OR message)
    query = request.GET.get('q')
    if query:
        announcements = announcements.filter(
            Q(title__icontains=query) |
            Q(message__icontains=query)
        )

    # ✏️ INLINE EDIT (GET)
    edit_id = request.GET.get('edit')

    # 🗑 INLINE DELETE (GET)
    delete_id = request.GET.get('delete')
    if delete_id:
        announcement = get_object_or_404(Announcement, id=delete_id)
        announcement.delete()
        messages.success(request, "Announcement deleted successfully")
        return redirect(request.path)

    # 💾 SAVE EDIT (POST)
    if request.method == 'POST':
        announcement_id = request.POST.get('announcement_id')
        announcement = get_object_or_404(Announcement, id=announcement_id)

        announcement.title = request.POST.get('title')
        announcement.message = request.POST.get('message')
        announcement.save()

        messages.success(request, "Announcement updated successfully")
        return redirect(request.path)

    return render(request, 'announcement_list.html', {
        'announcements': announcements,
        'edit_id': edit_id
    })

